from bs4 import BeautifulSoup
import requests
import time
import json
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.chart import LineChart, Reference

india = {
    'Active Cases': 0,
    'Total Cases': 0,
    'Cured': 0,
    'Deaths': 0,
    'Migrated': 0
    }

mah = {
    'Active Cases': 0,
    'Total Cases': 0,
    'Cured': 0,
    'Deaths': 0
    }
mah_dist = {
    "Mumnai": {
        'Total Cases': 0,
        'Active Cases': 0,
        'Cured': 0,
        'Deaths': 0,
        "delta":{
            "Total Cases": 0,
            'Cured': 0,
            'Deaths': 0
            }
        },
    "Pune": {
        'Total Cases': 0,
        'Active Cases': 0,
        'Cured': 0,
        'Deaths': 0,
        "delta":{
            "Total Cases": 0,
            'Cured': 0,
            'Deaths': 0
            }
        },
    "Thane": {
        'Total Cases': 0,
        'Active Cases': 0,
        'Cured': 0,
        'Deaths': 0,
        "delta":{
            "Total Cases": 0,
            'Cured': 0,
            'Deaths': 0
            }
        }
    }
source = requests.get('https://www.mohfw.gov.in/').text
soup = BeautifulSoup(source,'lxml')
active = soup.find('li', class_ = 'bg-blue')
cured = soup.find('li', class_ = 'bg-green')
deaths = soup.find('li', class_ = 'bg-red')
migrated = soup.find('li', class_ = 'bg-orange')
india['Active Cases'] = active.strong.text
india['Cured'] = cured.strong.text
india['Deaths'] = deaths.strong.text
india['Migrated'] = migrated.strong.text

extract_contents = lambda row: [x.text.replace('\n','') for x in row]
stats = []
all_rows = soup.find_all('tr')

for row in all_rows:
    stat = extract_contents(row.find_all('td'))
    if len(stat) == 5:
        stats.append(stat)

for item in stats:
    india['Total Cases'] += int(item[2])
    if item[1] == "Maharashtra":
        mah['Total Cases'] = item[2]
        mah['Active Cases'] = int(item[2]) - int(item[3]) - int(item[4])
        mah['Cured'] = item[3]
        mah['Deaths'] = item[4]
    #print(item[0] + ". " + item[1] + " - " + item[2] + ", " + item[3] + ", " + item[4]) 
#print(india)

state = soup.find('table', class_ = 'table table-striped')
#print(state.tbody.text)



response = requests.get("https://api.covid19india.org/state_district_wise.json")
time.sleep(2)
state_data = response.json()
time.sleep(2)


#print("India: ")
#print(india)
#print("Maharashtra: ")
#print(mah)

district = ["Mumbai", "Pune", "Thane"]
for d in district:
    if d in state_data["Maharashtra"]["districtData"]:
        mah_dist[d] = state_data["Maharashtra"]["districtData"][d]
        del mah_dist[d]['notes']
        del mah_dist[d]['delta']
        #print(d + ": ")
        #print(mah_dist[d]) 
        
region = [india, mah, mah_dist["Mumbai"], mah_dist["Pune"], mah_dist["Thane"]]
header = ["Date", "Total Cases", "Active Cases", "Recovered", "Deaths"]
data = []

for place in region:
    today = date.today()
    day = "{}-{}-{}".format(today.day, today.month, today.year)
    lst = [day]
    j=0
    for cases, numbers in place.items():
        lst.append(numbers)
        j+=1
        if j==4:
            break
    data.append(lst)
    j=0
    lst=[]
#print(header)    
print(data)

filename = "covid19.xlsx"    
workbook = load_workbook(filename)
india_sheet = workbook["India"]
mah_sheet = workbook["Maharashtra"]
mumbai_sheet = workbook["Mumbai"]
pune_sheet = workbook["Pune"]
thane_sheet = workbook["Thane"]
#print(workbook.sheetnames)
all_sheets = [india_sheet, mah_sheet, mumbai_sheet, pune_sheet, thane_sheet]


def update_sheet(data):
    
    for sheet in all_sheets:
        for row in data:
            if all_sheets.index(sheet) == data.index(row):
                sheet.append(row)
        
def print_rows():
    for sheet in all_sheets:
        print(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            print(row)


update_sheet(data)
print_rows()

##chart = LineChart()
##data = Reference(worksheet=india_sheet,
##                 min_row=2,
##                 max_row=4,
##                 min_col=1,
##                 max_col=5)
##
##chart.add_data(data, titles_from_data=True)
##india_sheet.add_chart(chart, "G3")
workbook.save(filename)
print("DONE")        






