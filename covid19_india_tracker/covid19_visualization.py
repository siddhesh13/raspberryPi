from bs4 import BeautifulSoup
import requests
import time
import json
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.chart import LineChart, Reference
import pandas as pd
import matplotlib.pyplot as plt
#import geopandas as gpd
from prettytable import PrettyTable
import numpy as np
india = {
    'Active Cases': 0,
    'Total Cases': 0,
    'Cured': 0,
    'Deaths': 0,
    'Migrated': 0
    }

mah = {
    'Active Cases': 0,
    'Total Cases': 0,
    'Cured': 0,
    'Deaths': 0
    }
mah_dist = {
    "Mumnai": {
        'Total Cases': 0,
        'Active Cases': 0,
        'Cured': 0,
        'Deaths': 0,
        "delta":{
            "Total Cases": 0,
            'Cured': 0,
            'Deaths': 0
            }
        },
    "Pune": {
        'Total Cases': 0,
        'Active Cases': 0,
        'Cured': 0,
        'Deaths': 0,
        "delta":{
            "Total Cases": 0,
            'Cured': 0,
            'Deaths': 0
            }
        },
    "Thane": {
        'Total Cases': 0,
        'Active Cases': 0,
        'Cured': 0,
        'Deaths': 0,
        "delta":{
            "Total Cases": 0,
            'Cured': 0,
            'Deaths': 0
            }
        }
    }
source = requests.get('https://www.mohfw.gov.in/').text
soup = BeautifulSoup(source,'lxml')
active = soup.find('li', class_ = 'bg-blue')
cured = soup.find('li', class_ = 'bg-green')
deaths = soup.find('li', class_ = 'bg-red')
migrated = soup.find('li', class_ = 'bg-orange')
india['Active Cases'] = active.strong.text
india['Cured'] = cured.strong.text
india['Deaths'] = deaths.strong.text
india['Migrated'] = migrated.strong.text

extract_contents = lambda row: [x.text.replace('\n','') for x in row]
stats = []
all_rows = soup.find_all('tr')

for row in all_rows:
    stat = extract_contents(row.find_all('td'))
    if len(stat) == 5:
        stats.append(stat)

for item in stats:
    india['Total Cases'] += int(item[2])
    if item[1] == "Maharashtra":
        mah['Total Cases'] = item[2]
        mah['Active Cases'] = int(item[2]) - int(item[3]) - int(item[4])
        mah['Cured'] = item[3]
        mah['Deaths'] = item[4]
    #print(item[0] + ". " + item[1] + " - " + item[2] + ", " + item[3] + ", " + item[4]) 
#print(india)

state = soup.find('table', class_ = 'table table-striped')
#print(state.tbody.text)



response = requests.get("https://api.covid19india.org/state_district_wise.json")
time.sleep(2)
state_data = response.json()
time.sleep(2)


#print("India: ")
#print(india)
#print("Maharashtra: ")
#print(mah)

district = ["Mumbai", "Pune", "Thane"]
for d in district:
    if d in state_data["Maharashtra"]["districtData"]:
        mah_dist[d] = state_data["Maharashtra"]["districtData"][d]
        del mah_dist[d]['notes']
        del mah_dist[d]['delta']
        #print(d + ": ")
        #print(mah_dist[d]) 
        
region = [india, mah, mah_dist["Mumbai"], mah_dist["Pune"], mah_dist["Thane"]]
header = ["Date", "Total Cases", "Active Cases", "Recovered", "Deaths"]
data = []

for place in region:
    today = date.today()
    day = "{}-{}-{}".format(today.day, today.month, today.year)
    lst = [day]
    j=0
    for cases, numbers in place.items():
        lst.append(numbers)
        j+=1
        if j==4:
            break
    data.append(lst)
    j=0
    lst=[]
#print(header)    
print(data)

filename = "covid19.xlsx"    
workbook = load_workbook(filename)
india_sheet = workbook["India"]
mah_sheet = workbook["Maharashtra"]
mumbai_sheet = workbook["Mumbai"]
pune_sheet = workbook["Pune"]
thane_sheet = workbook["Thane"]
#print(workbook.sheetnames)
all_sheets = [india_sheet, mah_sheet, mumbai_sheet, pune_sheet, thane_sheet]


def update_sheet(data):
    
    for sheet in all_sheets:
        for row in data:
            if all_sheets.index(sheet) == data.index(row):
                sheet.append(row)
all_data=[]        
def print_rows():
    for sheet in all_sheets:
        print(sheet.title)
        table = PrettyTable()
        table.field_names = (header)
        i=0
        for row in sheet.iter_rows(values_only=True):
            #print(row)
            #prettytable representation
            if i != 0:
                table.add_row(row)
            i+=1    
        print(table)    
            
line_chart_data=[]
def get_cols():
    for sheet in all_sheets:
        for col in sheet.iter_cols(min_row=2,values_only=True):
            line_chart_data.append(col)
get_cols()
#print(line_chart_data)            

#update_sheet(data)
print_rows()

workbook.save(filename)
print("DONE")        

#matplotlib donut
# donut chart representing nationwide total confirmed, cured and deceased cases
group_size = [india['Total Cases'], 
              india['Cured'], 
              india['Deaths']]

group_labels = ['Confirmed\n' + str(india['Active Cases']), 
                'Recovered\n' + str(india['Cured']), 
                'Deceased\n'  + str(india['Deaths'])]
custom_colors = ['skyblue','yellowgreen','tomato']

plt.figure(figsize = (5,5))
plt.pie(group_size, labels = group_labels, colors = custom_colors)
central_circle = plt.Circle((0,0), 0.5, color = 'white')
fig = plt.gcf()
fig.gca().add_artist(central_circle)
plt.rc('font', size = 12) 
plt.title('Nationwide total Confirmed, Recovered and Deceased Cases', fontsize = 16)
plt.show()


def list_items_to_int(lst):
    temp_list = []
    for i in lst:
        temp_list.append(int(i))
    return temp_list    



lst_num=[6,7,8,9]
#daily chart
def daily_value(col_lst, colm):
    daily_line_chart=[]
    for lst in col_lst:
        if col_lst.index(lst) == colm:
            daily_line_chart=[0]
            print(lst)
            for num in lst:
                if lst.index(num) != (len(lst)-1): 
                    daily_line_chart.append(int(lst[lst.index(num)+1]) - int(num))
    return daily_line_chart    


#Maharashtra graph
index = np.arange(len(line_chart_data[5]))
#indexY = np.arange(len(daily_line_chart))
plt.bar(index, daily_value(line_chart_data,7))
plt.xlabel('Date', fontsize=5)
plt.ylabel('Number of daily cases', fontsize=5)
plt.xticks(index, line_chart_data[5], fontsize=5, rotation=45)
plt.title('Maharashtra Covid19 Tracker - daily cases')
plt.show()  

#line chart

plt.plot(line_chart_data[5],list_items_to_int(line_chart_data[7]), label="Total Cases")

plt.plot(line_chart_data[5],list_items_to_int(line_chart_data[6]), label="Active Cases")

##plt.plot(line_chart_data[5],list_items_to_int(line_chart_data[8]), label="Recovered")
##
##plt.plot(line_chart_data[5],list_items_to_int(line_chart_data[9]), label="Deaths")
plt.xlabel('Date')
plt.xticks(rotation=45)
plt.ylabel('Number of cases')
# Set a title of the current axes.
plt.title('Maharashtra Covid19 Tracker')
# show a legend on the plot
plt.legend(loc="upper left", title="Legend", frameon=True)
plt.show()


#pune, thane graph
#line chart
fig, ax = plt.subplots(2,2, figsize=(10,10))

fig.suptitle("Pune, Thane covid19 tracker")

ax[0, 0].set_title('Total Cases')
ax[0,0].plot(line_chart_data[10],list_items_to_int(line_chart_data[17]), color="blue", label="Pune")
ax[0,0].plot(line_chart_data[10],list_items_to_int(line_chart_data[22]), color="green", label="Thane")

ax[0, 1].set_title('Active Cases')
ax[0,1].plot(line_chart_data[10],list_items_to_int(line_chart_data[16]), color="blue", label="Pune")
ax[0,1].plot(line_chart_data[10],list_items_to_int(line_chart_data[21]), color="green", label="Thane")

ax[1, 0].set_title('Deaths')
ax[1,0].plot(line_chart_data[10],list_items_to_int(line_chart_data[18]), color="blue", label="Pune")
ax[1,0].plot(line_chart_data[10],list_items_to_int(line_chart_data[23]), color="green", label="Thane")

ax[1, 1].set_title('Recovered')
ax[1,1].plot(line_chart_data[10],list_items_to_int(line_chart_data[19]), color="blue", label="Pune")
ax[1,1].plot(line_chart_data[10],list_items_to_int(line_chart_data[24]), color="green", label="Thane")

plt.legend(loc="lower right", title="Legend", frameon=True)
plt.show()



#Mumbai graph
index = np.arange(len(line_chart_data[10]))
plt.bar(index, daily_value(line_chart_data,12))
plt.xlabel('Date', fontsize=5)
plt.ylabel('Number of daily cases', fontsize=5)
plt.xticks(index, line_chart_data[10], fontsize=5, rotation=45)
plt.title('Mumbai Covid19 Tracker - daily cases')
plt.show()  
print(daily_value(line_chart_data,12))
#line chart

plt.plot(line_chart_data[10],list_items_to_int(line_chart_data[12]), label="Total Cases")

plt.plot(line_chart_data[10],list_items_to_int(line_chart_data[11]), label="Active Cases")

plt.xlabel('Date')
plt.xticks(rotation=45)
plt.ylabel('Number of cases')
# Set a title of the current axes.
plt.title('Mumbai Covid19 Tracker')
# show a legend on the plot
plt.legend(loc="upper left", title="Legend", frameon=True)
plt.show()
